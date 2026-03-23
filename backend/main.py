import time
from datetime import datetime
import pandas as pd
from flask import Flask, jsonify, request # Added jsonify and request
from flask_socketio import SocketIO
from flask_cors import CORS
import os
import joblib # Added for ML model
from threading import Thread
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# --- 1. PATH SETUP ---
backend_dir = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.abspath(os.path.join(backend_dir, "..", "data_science", "sensors.xlsx"))
MODEL_PATH = os.path.join(backend_dir, 'crop_risk_model.pkl')

# Load the AI Brain
try:
    model = joblib.load(MODEL_PATH)
    print("🧠 AI Brain loaded successfully.")
except:
    print("⚠️ Warning: .pkl model not found. Retrain the model first.")

# --- 2. PARAMETRIC ENGINE ---
def calculate_parametric_metrics(t, m, h, r):
    if t > 35: r_t = max(0, (t - 35) / (50 - 35) * 100)
    elif t < 18: r_t = max(0, (18 - t) / (18 - 5) * 100)
    else: r_t = 0
    
    if m < 40: r_m = max(0, (40 - m) / (40 - 20) * 100)
    elif m > 80: r_m = max(0, (m - 80) / (95 - 80) * 100)
    else: r_m = 0
    
    r_h = max(0, (h - 75) / (95 - 75) * 100) if h > 75 else 0
    r_r = max(0, (r - 30) / (60 - 30) * 100) if r > 30 else 0

    total_risk = min(100, max(r_t, r_m, r_h, r_r))
    drivers = {"Temperature": r_t, "Soil Moisture": r_m, "Humidity": r_h, "Rainfall": r_r}
    primary_driver = max(drivers, key=drivers.get) if total_risk > 0 else "Optimal"

    payout_pct = 0
    if total_risk > 40:
        payout_pct = min(100, (total_risk - 40) / (90 - 40) * 100)

    return round(total_risk, 2), primary_driver, round(payout_pct, 2)

# --- 3. AI PREDICTION ROUTE ---
@app.route('/predict', methods=['POST'])
def predict():
    try:
        data = request.json
        # Convert incoming JSON to list for model
        features = [[
            data['temperature'], 
            data['soil_moisture'], 
            data['humidity'], 
            data['rainfall']
        ]]
        prediction = model.predict(features)[0]
        return jsonify({"risk_score": round(prediction, 2)})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# --- 4. DATA SYNC ---
def get_full_history():
    if not os.path.exists(EXCEL_PATH): return []
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        if not ws.cell(row=1, column=5).value: ws.cell(row=1, column=5).value = "timestamp"
        updated = False
        for row in range(2, 1000): 
            temp_val = ws.cell(row=row, column=1).value
            if temp_val is None: break
            ts_cell = ws.cell(row=row, column=5)
            if ts_cell.value is None or str(ts_cell.value).strip() in ["", "nan", "None"]:
                ts_cell.value = datetime.now().strftime("%d %b, %I:%M %p")
                updated = True
        if updated:
            wb.save(EXCEL_PATH)
            print("💾 SUCCESS: Timestamps permanently saved.")
        wb.close()
    except PermissionError: print("🕒 Excel is open. Syncing data live.")
    except Exception as e: print(f"🕒 Timestamp error: {e}")

    try:
        df = pd.read_excel(EXCEL_PATH, engine='openpyxl')
        if df.empty: return []
        history = []
        for i in range(len(df)):
            try:
                t, m, h, r = float(df.iloc[i, 0]), float(df.iloc[i, 1]), float(df.iloc[i, 2]), float(df.iloc[i, 3])
                file_ts = df.iloc[i, 4] if df.shape[1] > 4 else None
                ts = str(file_ts) if pd.notnull(file_ts) and str(file_ts).strip() not in ["", "nan", "None"] else datetime.now().strftime("%I:%M %p") + " (Live)"
                risk, driver, payout = calculate_parametric_metrics(t, m, h, r)
                history.append({
                    "temperature": t, "soil_moisture": m, "humidity": h, "rainfall": r,
                    "timestamp": ts, "risk_score": risk, "health_score": round(100 - risk, 2),
                    "primary_driver": driver, "payout_pct": payout, "sensor_id": "PICT-NODE-001"
                })
            except: continue
        return history
    except Exception as e:
        print(f"❌ Read Error: {e}")
        return []

@socketio.on('connect')
def handle_connect():
    history = get_full_history()
    socketio.emit('initial_data_history', history)

def monitor_file():
    last_mtime = 0
    while True:
        try:
            if os.path.exists(EXCEL_PATH):
                current_mtime = os.path.getmtime(EXCEL_PATH)
                if current_mtime != last_mtime:
                    time.sleep(0.5) 
                    history = get_full_history()
                    socketio.emit('initial_data_history', history)
                    last_mtime = current_mtime
        except: pass
        time.sleep(1)

if __name__ == "__main__":
    Thread(target=monitor_file, daemon=True).start()
    socketio.run(app, port=5000, debug=False, allow_unsafe_werkzeug=True)